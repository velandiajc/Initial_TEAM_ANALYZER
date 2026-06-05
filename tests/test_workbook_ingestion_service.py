import pandas as pd
import pytest
from openpyxl import Workbook

from app.services.workbook_ingestion_service import WorkbookIngestionService


def create_valid_workbook(path):
    with pd.ExcelWriter(path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "Agent Name": ["Agent One", "Agent Two"],
                "Agent ID": ["A001", "A002"],
                "OSAT": [9, 8],
                "CSAT": [95, 88],
                "QA": [97, 91],
            }
        ).to_excel(writer, sheet_name="Surveys", index=False)

        pd.DataFrame(
            {
                "Metric": ["Volume", "AHT"],
                "Value": [20, 310],
            }
        ).to_excel(writer, sheet_name="Performance", index=False)


def create_problem_workbook(path):
    workbook = Workbook()

    surveys = workbook.active
    surveys.title = "Surveys"
    surveys.append(["Agent Name", "Agent ID", "OSAT", "CSAT", "QA", "OSAT"])
    surveys.append(["Agent One", "A001", -1, 90, 95, 9])
    surveys.append(["Agent Two", "A002", 13, 120, 105, 8])
    surveys.append([None, "A003", 8, 85, 90, 8])
    surveys.append([None, None, None, None, None, None])
    surveys.append(["Agent Four", "A004", 7, 80, 85, 7])

    empty = workbook.create_sheet("Empty Sheet")
    empty.append(["Agent Name", "OSAT"])

    workbook.save(path)


def create_delayed_header_workbook(path):
    workbook = Workbook()

    roaster = workbook.active
    roaster.title = "Roaster"
    roaster.append(["Roaster Operational Report"])
    roaster.append([])
    roaster.append([])
    roaster.append(["Generated for supervisor review"])
    roaster.append(["Agent Name", "ID", "Supervisor", "Status", "Schedule"])
    roaster.append(["Agent One", "A001", "Supervisor One", "Active", "Morning"])
    roaster.append(["Agent Two", "A002", "Supervisor Two", "Leave", "Evening"])

    qa_work = workbook.create_sheet("QA_WORK")
    qa_work.append([])
    qa_work.append([])
    qa_work.append(["QA Work Queue"])
    qa_work.append([])
    qa_work.append(["Agent Name", "Contact ID", "QA", "Disposition"])
    qa_work.append(["Agent One", "C001", 95, "Complete"])

    workbook.save(path)


def issue_types(validation):
    return [issue.issue_type for issue in validation.issues]


def test_valid_workbook_with_multiple_sheets(tmp_path):
    workbook_path = tmp_path / "valid.xlsx"
    create_valid_workbook(workbook_path)

    result = WorkbookIngestionService().ingest(workbook_path)

    assert result.inventory.sheet_count == 2
    assert [sheet.sheet_name for sheet in result.inventory.sheets] == [
        "Surveys",
        "Performance",
    ]
    assert result.validation.issue_count == 0


def test_inventory_captures_sheet_shape_columns_and_samples(tmp_path):
    workbook_path = tmp_path / "valid.xlsx"
    create_valid_workbook(workbook_path)

    inventory = WorkbookIngestionService().build_inventory(workbook_path)
    survey_sheet = inventory.sheets[0]

    assert survey_sheet.row_count == 2
    assert survey_sheet.column_count == 5
    assert survey_sheet.column_names == [
        "Agent Name",
        "Agent ID",
        "OSAT",
        "CSAT",
        "QA",
    ]
    assert survey_sheet.sample_rows[0]["Agent Name"] == "Agent One"


def test_detects_header_on_row_five(tmp_path):
    workbook_path = tmp_path / "delayed_header.xlsx"
    create_delayed_header_workbook(workbook_path)

    inventory = WorkbookIngestionService().build_inventory(workbook_path)
    roaster_sheet = inventory.sheets[0]

    assert roaster_sheet.header_row_number == 5
    assert roaster_sheet.column_names == [
        "Agent Name",
        "ID",
        "Supervisor",
        "Status",
        "Schedule",
    ]
    assert roaster_sheet.row_count == 2


def test_blank_rows_before_header_are_not_validation_issues(tmp_path):
    workbook_path = tmp_path / "delayed_header.xlsx"
    create_delayed_header_workbook(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    blank_issues = [
        issue
        for issue in validation.issues
        if issue.issue_type == "blank_row"
    ]

    assert blank_issues == []


def test_title_row_before_header_is_not_used_as_header(tmp_path):
    workbook_path = tmp_path / "delayed_header.xlsx"
    create_delayed_header_workbook(workbook_path)

    inventory = WorkbookIngestionService().build_inventory(workbook_path)
    roaster_sheet = inventory.sheets[0]

    assert roaster_sheet.header_row_number == 5
    assert "Roaster Operational Report" not in roaster_sheet.column_names
    assert roaster_sheet.sample_rows[0]["Agent Name"] == "Agent One"


def test_duplicate_blank_column_names_do_not_count_as_duplicate_column(tmp_path):
    workbook_path = tmp_path / "blank_headers.xlsx"
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Blank Headers"
    sheet.append(["Agent Name", None, None, "OSAT"])
    sheet.append(["Agent One", "unused", "unused", 9])
    workbook.save(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    assert not any(
        issue.issue_type == "duplicate_column"
        for issue in validation.issues
    )


def test_empty_sheet_is_flagged(tmp_path):
    workbook_path = tmp_path / "problems.xlsx"
    create_problem_workbook(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    assert "empty_sheet" in issue_types(validation)


def test_duplicate_columns_are_flagged(tmp_path):
    workbook_path = tmp_path / "problems.xlsx"
    create_problem_workbook(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    duplicates = [
        issue
        for issue in validation.issues
        if issue.issue_type == "duplicate_column"
    ]

    assert duplicates
    assert duplicates[0].column_name == "osat"


def test_osat_below_zero_is_flagged(tmp_path):
    workbook_path = tmp_path / "problems.xlsx"
    create_problem_workbook(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    assert any(
        issue.issue_type == "negative_numeric_value"
        and issue.column_name == "OSAT"
        and issue.value == "-1"
        for issue in validation.issues
    )


def test_osat_above_ten_is_flagged(tmp_path):
    workbook_path = tmp_path / "problems.xlsx"
    create_problem_workbook(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    assert any(
        issue.issue_type == "osat_above_10"
        and issue.column_name == "OSAT"
        and issue.value == "13"
        for issue in validation.issues
    )


def test_csat_above_one_hundred_is_flagged(tmp_path):
    workbook_path = tmp_path / "problems.xlsx"
    create_problem_workbook(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    assert any(
        issue.issue_type == "csat_above_100"
        and issue.column_name == "CSAT"
        and issue.value == "120"
        for issue in validation.issues
    )


def test_qa_above_one_hundred_is_flagged(tmp_path):
    workbook_path = tmp_path / "problems.xlsx"
    create_problem_workbook(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    assert any(
        issue.issue_type == "qa_above_100"
        and issue.column_name == "QA"
        and issue.value == "105"
        for issue in validation.issues
    )


def test_missing_agent_name_is_flagged(tmp_path):
    workbook_path = tmp_path / "problems.xlsx"
    create_problem_workbook(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    assert any(
        issue.issue_type == "missing_agent_identifier"
        and issue.column_name == "Agent Name"
        for issue in validation.issues
    )


def test_blank_rows_are_flagged(tmp_path):
    workbook_path = tmp_path / "problems.xlsx"
    create_problem_workbook(workbook_path)

    validation = WorkbookIngestionService().validate_workbook(workbook_path)

    assert "blank_row" in issue_types(validation)


def test_sample_rows_convert_nan_and_truncate_long_values(tmp_path):
    workbook_path = tmp_path / "samples.xlsx"
    long_value = "x" * 150

    with pd.ExcelWriter(workbook_path, engine="openpyxl") as writer:
        pd.DataFrame(
            {
                "Agent Name": [
                    "Agent One",
                    "Agent Two",
                    "Agent Three",
                    "Agent Four",
                    "Agent Five",
                    "Agent Six",
                ],
                "CSAT": [95, None, 88, 87, 86, 85],
                "Comment": [long_value, "Short", "Third", "Fourth", "Fifth", "Sixth"],
            }
        ).to_excel(writer, sheet_name="Samples", index=False)

    inventory = WorkbookIngestionService().build_inventory(workbook_path)
    sample_rows = inventory.sheets[0].sample_rows

    assert len(sample_rows) == 5
    assert sample_rows[1]["CSAT"] == ""
    assert sample_rows[0]["Comment"] == "x" * 120


def test_missing_workbook_raises_file_not_found(tmp_path):
    with pytest.raises(FileNotFoundError, match="Workbook file not found"):
        WorkbookIngestionService().ingest(tmp_path / "missing.xlsx")


def test_non_xlsx_input_raises_value_error(tmp_path):
    text_file = tmp_path / "workbook.xls"
    text_file.write_text("not xlsx", encoding="utf-8")

    with pytest.raises(ValueError, match="only supports .xlsx"):
        WorkbookIngestionService().ingest(text_file)


def test_renderers_generate_inventory_and_validation_markdown(tmp_path):
    workbook_path = tmp_path / "problems.xlsx"
    create_problem_workbook(workbook_path)
    service = WorkbookIngestionService()
    result = service.ingest(workbook_path)

    inventory_markdown = service.render_inventory_markdown(result.inventory)
    validation_markdown = service.render_validation_markdown(result.validation)

    assert "# Workbook Inventory" in inventory_markdown
    assert "## Sheet: Surveys" in inventory_markdown
    assert "- Detected header row: 1" in inventory_markdown
    assert "| Agent Name | Agent ID | OSAT | CSAT | QA | OSAT |" in inventory_markdown
    assert "# Workbook Validation" in validation_markdown
    assert "osat_above_10" in validation_markdown
    assert "No validation issues found." not in validation_markdown
