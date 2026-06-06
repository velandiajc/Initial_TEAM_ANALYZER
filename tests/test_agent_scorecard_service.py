from openpyxl import Workbook

from app.services.agent_scorecard_service import AgentScorecardService


def create_roaster_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roaster"
    sheet.append(["Roaster Performance Snapshot"])
    sheet.append([])
    sheet.append([])
    sheet.append(["Generated for Sprint 1 demo"])
    sheet.append(
        [
            "Name",
            "ID",
            "Email",
            "Supervisor",
            "CSAT MTD",
            "Detractors",
            "Open DSAT",
            "QA MTD",
            "Main DSAT Driver",
            "QA Main Gap",
            "Risk_Final",
            "Coaching Focus",
            "Coaching Needed",
            "Coaching Action (Auto)",
            "Coaching Topic (Auto)",
            "Coaching Reason (Auto)",
            "Critical Flag",
            "Schedule",
        ]
    )
    sheet.append(
        [
            "Daniela Talero",
            "A001",
            "daniela@example.com",
            "Supervisor One",
            72.22,
            0,
            0,
            0.8144,
            "",
            "",
            "Low",
            "Maintain empathy",
            "No",
            "Monitor",
            "CX basics",
            "Strong current results",
            "No",
            "Morning",
        ]
    )
    sheet.append([None for _ in range(18)])
    sheet.append(
        [
            "Nicole Molina",
            "A002",
            "nicole@example.com",
            "Supervisor Two",
            71,
            3,
            2,
            82,
            "Resolution",
            "Documentation",
            "High",
            "Ownership",
            "Yes",
            "Review DSAT calls",
            "Resolution quality",
            "Open DSAT volume",
            "Yes",
            "Evening",
        ]
    )
    workbook.save(path)


def create_minimal_roaster_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roaster"
    sheet.append(["Name", "ID"])
    sheet.append(["Cristian Pinto", "A003"])
    workbook.save(path)


def create_risk_roaster_workbook(path, risk_level, critical_flag="No"):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roaster"
    sheet.append(
        [
            "Name",
            "ID",
            "CSAT MTD",
            "QA MTD",
            "Detractors",
            "Open DSAT",
            "Main DSAT Driver",
            "QA Main Gap",
            "Risk_Final",
            "Coaching Action (Auto)",
            "Coaching Reason (Auto)",
            "Critical Flag",
        ]
    )
    sheet.append(
        [
            "Risk Agent",
            "A100",
            66.5,
            78,
            3,
            2,
            "Resolution",
            "Documentation",
            risk_level,
            "Review DSAT calls",
            "Open DSAT trend",
            critical_flag,
        ]
    )
    workbook.save(path)


def create_mojibake_roaster_workbook(path):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = "Roaster"
    sheet.append(
        [
            "Name",
            "ID",
            "CSAT MTD",
            "QA MTD",
            "Coaching Reason (Auto)",
        ]
    )
    sheet.append(
        [
            "Analyn Brice\u00c3\u00b1o",
            "A200",
            88,
            91,
            "\u00f0\u0178\u201d\u00a5 DSAT spike \u00e2\u20ac\u201d needs review",
        ]
    )
    workbook.save(path)


def test_builds_scorecards_from_roaster_with_detected_header(tmp_path):
    workbook_path = tmp_path / "team.xlsx"
    create_roaster_workbook(workbook_path)

    report = AgentScorecardService().build_report(workbook_path)

    assert report.sheet_name == "Roaster"
    assert len(report.scorecards) == 2
    assert report.scorecards[0].agent_name == "Daniela Talero"
    assert report.scorecards[0].csat_mtd == "72.22%"
    assert report.scorecards[0].qa_mtd == "81.44%"
    assert report.scorecards[1].agent_name == "Nicole Molina"
    assert report.scorecards[1].critical_flag == "Yes"


def test_skips_blank_rows(tmp_path):
    workbook_path = tmp_path / "team.xlsx"
    create_roaster_workbook(workbook_path)

    report = AgentScorecardService().build_report(workbook_path)

    assert [
        scorecard.agent_name
        for scorecard in report.scorecards
    ] == [
        "Daniela Talero",
        "Nicole Molina",
    ]


def test_handles_missing_columns_gracefully(tmp_path):
    workbook_path = tmp_path / "minimal.xlsx"
    create_minimal_roaster_workbook(workbook_path)

    report = AgentScorecardService().build_report(workbook_path)
    scorecard = report.scorecards[0]
    markdown = AgentScorecardService().render_markdown(report)

    assert scorecard.agent_name == "Cristian Pinto"
    assert scorecard.email == ""
    assert scorecard.csat_mtd == ""
    assert "| Email | Not available |" in markdown
    assert "| CSAT MTD | Not available |" in markdown


def test_formats_percentage_values(tmp_path):
    workbook_path = tmp_path / "team.xlsx"
    create_roaster_workbook(workbook_path)

    report = AgentScorecardService().build_report(workbook_path)
    scorecard = report.scorecards[1]

    assert scorecard.csat_mtd == "71.00%"
    assert scorecard.qa_mtd == "82.00%"


def test_converts_qa_decimal_to_percentage(tmp_path):
    workbook_path = tmp_path / "team.xlsx"
    create_roaster_workbook(workbook_path)

    report = AgentScorecardService().build_report(workbook_path)

    assert report.scorecards[0].qa_mtd == "81.44%"


def test_cleans_mojibake_text(tmp_path):
    workbook_path = tmp_path / "mojibake.xlsx"
    create_mojibake_roaster_workbook(workbook_path)

    report = AgentScorecardService().build_report(workbook_path)
    scorecard = report.scorecards[0]

    assert scorecard.agent_name == "Analyn Briceño"
    assert "—" in scorecard.coaching_reason
    assert "ðŸ”¥" not in scorecard.coaching_reason
    assert "high urgency" in scorecard.coaching_reason


def test_generates_supervisor_recommendation_for_high_or_critical_risk(tmp_path):
    workbook_path = tmp_path / "critical.xlsx"
    create_risk_roaster_workbook(
        workbook_path,
        risk_level="Critical"
    )

    report = AgentScorecardService().build_report(workbook_path)
    recommendation = report.scorecards[0].supervisor_recommendation

    assert "Prioritize a risk review this week" in recommendation
    assert "Resolution" in recommendation
    assert "Documentation" in recommendation
    assert "Review DSAT calls" in recommendation


def test_generates_supervisor_recommendation_for_moderate_risk(tmp_path):
    workbook_path = tmp_path / "moderate.xlsx"
    create_risk_roaster_workbook(
        workbook_path,
        risk_level="Moderate"
    )

    report = AgentScorecardService().build_report(workbook_path)
    recommendation = report.scorecards[0].supervisor_recommendation

    assert "targeted coaching" in recommendation
    assert "Moderate" in recommendation
    assert "2 open DSAT and 3 detractors" in recommendation


def test_generates_supervisor_recommendation_from_critical_flag(tmp_path):
    workbook_path = tmp_path / "team.xlsx"
    create_roaster_workbook(workbook_path)

    report = AgentScorecardService().build_report(workbook_path)
    high_risk_scorecard = report.scorecards[1]

    assert "immediate supervisor follow-up" in (
        high_risk_scorecard.supervisor_recommendation
    )


def test_render_markdown_keeps_scorecards_readable(tmp_path):
    workbook_path = tmp_path / "team.xlsx"
    create_roaster_workbook(workbook_path)

    report = AgentScorecardService().build_report(workbook_path)
    markdown = AgentScorecardService().render_markdown(report)

    assert "# Agent Scorecards" in markdown
    assert "## Daniela Talero" in markdown
    assert "| Field | Value |" in markdown
    assert "| Supervisor recommendation |" in markdown


def test_write_report_creates_markdown_file(tmp_path):
    workbook_path = tmp_path / "team.xlsx"
    output_path = tmp_path / "Reports" / "Sprint_1_Demos" / "agent_scorecards.md"
    create_roaster_workbook(workbook_path)

    report = AgentScorecardService().write_report(
        workbook_path,
        output_path
    )

    markdown = output_path.read_text(encoding="utf-8")

    assert output_path.exists()
    assert len(report.scorecards) == 2
    assert "Daniela Talero" in markdown
    assert "Nicole Molina" in markdown
